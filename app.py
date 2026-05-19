"""
OFFERING BANNER GENERATOR
Run: python app.py
Open: http://localhost:5000
"""

from pathlib import Path
import os
import csv
from io import BytesIO

from flask import Flask, render_template, send_from_directory, request, jsonify, send_file


ROOT = Path(__file__).parent
app = Flask(
    __name__,
    template_folder=str(ROOT / "templates"),
    static_folder=str(ROOT / "static"),
)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/icon.jpg")
def icon_file():
    return send_from_directory(ROOT, "icon.jpg")

@app.route("/download-sample")
def download_sample():
    import openpyxl
    # Create an in-memory workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Write headers
    headers = ["theme", "platforms", "modes", "features"]
    ws.append(headers)
    
    # Write sample data
    rows = [
        ["basic", "app", "light", "100+ Live Classes, Free Notes"],
        ["infinity", "app,web", "dark", "Unlimited Mock Tests, 24/7 Doubt Solving"],
        ["pro", "web", "light,dark", "1-on-1 Mentorship, Personal Guide"]
    ]
    for r in rows:
        ws.append(r)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="bulk_offering_sample.xlsx", as_attachment=True)

@app.route("/upload-bulk", methods=["POST"])
def upload_bulk():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    banners = []
    try:
        if file.filename.endswith('.csv'):
            stream = file.stream.read().decode('utf-8').splitlines()
            reader = csv.DictReader(stream)
            for row in reader:
                banners.append({
                    "theme": str(row.get("theme", "basic")).strip().lower(),
                    "platforms": [p.strip().lower() for p in str(row.get("platforms", "app")).split(",") if p.strip()],
                    "modes": [m.strip().lower() for m in str(row.get("modes", "light")).split(",") if m.strip()],
                    "features": [f.strip() for f in str(row.get("features", "")).split(",") if f.strip()]
                })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if all(v is None for v in row_dict.values()):
                    continue
                banners.append({
                    "theme": str(row_dict.get("theme", "basic")).strip().lower(),
                    "platforms": [p.strip().lower() for p in str(row_dict.get("platforms", "app")).split(",") if p.strip()],
                    "modes": [m.strip().lower() for m in str(row_dict.get("modes", "light")).split(",") if m.strip()],
                    "features": [f.strip() for f in str(row_dict.get("features", "")).split(",") if f.strip()]
                })
                
        return jsonify({"success": True, "banners": banners})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("\n" + "=" * 54)
    print("  OFFERING BANNER GENERATOR")
    print("  URL: http://localhost:5000")
    print("=" * 54 + "\n")
    app.run(debug=True, host="0.0.0.0", port=5000)
